"""Sync roster + skill ratings from Softball Lineup.xlsx into data.json.

Reads `../Softball Lineup.xlsx` (two tabs: Roster, Skill Notes).
Rewrites `data.json`["roster"] in place, preserving all other sections.
Run from repo root: `python scripts/sync_roster.py`

NOTE: The Roster tab has a two-row header (row 1 = section title, row 2 = column names).
      The Extras sub-section starts at row 18 with its own sub-header at row 19.
      The script reads both sections explicitly rather than relying on a single header row.
"""
import json
from pathlib import Path
from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parent.parent
XLSX = REPO_ROOT.parent / "Softball Lineup.xlsx"
DATA_JSON = REPO_ROOT / "data.json"

# Extras: these players are listed in the Extras section of the Roster tab.
EXTRAS = {"Brandon Ballard", "Ezra Pierce", "Rick Staadt"}


def read_roster_tab(wb):
    """Return list of {name, nickname, jersey, type} from the Roster tab.

    The tab has a two-section layout:
      Row 1: section title ('Roster')
      Row 2: column headers (Name, Jersey Name, Jersey Number, Position)
      Rows 3-15: main players (skip 'Extra' placeholder at jersey '00')
      Row 16: 'Extra' placeholder row — skip it
      Row 17: blank
      Row 18: 'Extras' section title — skip
      Row 19: sub-header row for extras — skip
      Rows 20-22: extra players (Brandon Ballard, Ezra Pierce, Rick Staadt)
    """
    ws = wb["Roster"]
    # Row 2 has the real headers (1-indexed); row 1 is a title.
    header_row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    headers = {v: i for i, v in enumerate(header_row) if v}
    name_col = headers["Name"]
    nick_col = headers["Jersey Name"]
    jersey_col = headers["Jersey Number"]

    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        name = row[name_col]
        nickname = row[nick_col]
        jersey = row[jersey_col]

        # Stop at a fully blank row that is followed only by blank rows
        # (we handle the Extras section by reading ALL rows and filtering).
        if name is None:
            continue

        # Skip section-title / sub-header rows (non-player rows)
        # These are rows where name is a label like 'Extra', 'Extras', 'Name'.
        if str(name).strip() in ("Extra", "Extras", "Name"):
            continue

        # Determine type: extras are listed in the EXTRAS set
        player_type = "extra" if str(name).strip() in EXTRAS else "main"

        rows.append({
            "name": str(name).strip(),
            "nickname": str(nickname).strip() if nickname else "",
            "jersey": str(jersey).strip() if jersey is not None else "",
            "type": player_type,
        })

    return rows


def read_skills_tab(wb):
    """Return dict keyed by Name with {glove, arm, bat} integers (0 if blank).

    Row 1 is the header row (Name, Jersey Name, Glove Strength (1-10), ...).
    """
    ws = wb["Skill Notes"]
    header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    headers = {v: i for i, v in enumerate(header_row) if v}
    name_col = headers["Name"]
    glove_col = headers["Glove Strength (1-10)"]
    arm_col = headers["Arm Strength (1-10)"]
    bat_col = headers["Batting Strength (1-10)"]

    skills = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[name_col]
        if not name:
            continue
        skills[str(name).strip()] = {
            "glove": int(row[glove_col] or 0),
            "arm": int(row[arm_col] or 0),
            "bat": int(row[bat_col] or 0),
        }
    return skills


def build_roster():
    wb = load_workbook(XLSX, data_only=True)
    roster_rows = read_roster_tab(wb)
    skills = read_skills_tab(wb)

    for row in roster_rows:
        row["skills"] = skills.get(row["name"], {"glove": 0, "arm": 0, "bat": 0})

    return roster_rows


def main():
    roster = build_roster()
    main_count = sum(1 for r in roster if r["type"] == "main")
    extra_count = sum(1 for r in roster if r["type"] == "extra")

    if DATA_JSON.exists():
        data = json.loads(DATA_JSON.read_text(encoding="utf-8"))
    else:
        data = {}

    data["roster"] = roster
    DATA_JSON.write_text(json.dumps(data, indent=2) + "\n", encoding="utf-8")
    print(f"Wrote {len(roster)} players to {DATA_JSON}")
    print(f"  {main_count} main, {extra_count} extra")


if __name__ == "__main__":
    main()
